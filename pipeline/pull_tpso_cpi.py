#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pull_tpso_cpi.py — MEASURED headline Thai CPI/inflation, monthly (owner review 2026-08-02: "inflation
data is readily available. Shouldn't be using 2025." The Macro tab was showing a World Bank ANNUAL
average for calendar-year 2025, on a board of otherwise monthly/quarterly indicators).

WHO ACTUALLY PUBLISHES THAI CPI — NOT the Bank of Thailand
  Checked directly (2026-08-02): BOT's own "real sector" statistics page
  (bot.or.th/th/statistics/real-sector.html) does not host a CPI report at all. Its "ข้อมูลจาก
  กระทรวงพาณิชย์" (data from the Ministry of Commerce) panel links OUT to
  index.tpso.go.th/cpi — Thailand's headline CPI is compiled and published monthly by the Ministry of
  Commerce's Trade Policy and Strategy Office (สำนักงานนโยบายและยุทธศาสตร์การค้า, TPSO), not BOT. (GDP
  is likewise NESDC's, not BOT's — see pull_nesdc_gdp.py.) BOT's own statistics only cover things it
  itself administers: interest rates, FX, reserves, external debt, balance of payments, monetary
  aggregates. This puller goes to the primary compiler (TPSO) rather than mislabelling a Ministry of
  Commerce series as a BOT one.

PULL METHOD
  index.tpso.go.th is a Next.js SPA with no working public JSON API found (its /api/* routes serve
  only category metadata, not the time series — the actual chart data loads through a private
  backend proxy not reachable directly). But its own "cpi" page bundle
  (_next/static/chunks/pages/cpi-*.js) references a plain static download it uses for the
  "ดาวน์โหลด" affordance:
    https://uploads.tpso.go.th/economic/pdf/Th_CPI_COICOP2018.xlsx
  This is TPSO's own COICOP-2018-basis CPI workbook (sheets Index / MoM / YoY / AoA), refreshed in
  place monthly — the exact same figures TPSO's own site charts. If this URL 404s in future (TPSO
  renames the export), rediscover it: fetch https://index.tpso.go.th/cpi, find the
  pages/cpi-<hash>.js chunk link, grep it for "uploads.tpso.go.th".

  Only the whole-Kingdom ("ประเทศไทย") "Total" (รหัส 0000000) row is kept — the headline number the
  Macro tab needs. The regional rows (Bangkok+vicinity, Central, North, Northeast, South) and the
  category breakdown are in the source file but not carried here; add them if a future card needs them.

VERIFIED 2026-08-02 (live pull): latest published month is 2026-05, headline YoY +2.79% — a sharp
swing from the near-zero/negative readings of early 2026 (Jan -0.66%, Feb -0.88%, Mar -0.08%) once
Apr/May turned positive (+2.89% / +2.79%). This is NOT the number the Macro tab currently shows
(-0.13%, a World Bank ANNUAL average for 2025) — they disagree in both sign and month.

DETERMINISM: the live pull's XLSX is cached (gitignored) at source-data/.tpso_cpi_raw/; the distilled
source-data/tpso_cpi.json is committed. `--check` re-parses the cached raw OFFLINE and byte-compares.
No wall clock in the data — `pulled` comes only from --stamp.

Run:
  python3 pull_tpso_cpi.py --stamp 2026-08-02   # download + parse + write
  python3 pull_tpso_cpi.py                       # default --stamp = today
  python3 pull_tpso_cpi.py --check               # offline byte-reproduce from cached raw
"""
import argparse
import datetime
import json
import os
import sys
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

OUT = os.path.join(ROOT, "source-data", "tpso_cpi.json")
RAW_DIR = os.path.join(ROOT, "source-data", ".tpso_cpi_raw")
RAW_XLSX = os.path.join(RAW_DIR, "Th_CPI_COICOP2018.xlsx")

XLSX_URL = "https://uploads.tpso.go.th/economic/pdf/Th_CPI_COICOP2018.xlsx"
CPI_PAGE_URL = "https://index.tpso.go.th/cpi"
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"}

NATION = "ประเทศไทย"
TOTAL_CODE = "0000000"
MONTH_COLS = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
              "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
SHEETS = ["Index", "MoM", "YoY", "AoA"]

# Spot-verification anchors — months already finalized at pull time; a mismatch means the parser
# broke (sheet/row order changed), not that TPSO revised settled history.
ANCHORS = {
    ("YoY", "2567", "ม.ค."): -1.11,
    ("YoY", "2569", "ม.ค."): -0.66,
    ("YoY", "2569", "พ.ค."): 2.79,
    ("Index", "2566", "ก.พ."): 100.25,
}


def _fetch(url, timeout=60):
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def _parse_xlsx(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = {}
    for sheet in SHEETS:
        ws = wb[sheet]
        by_year = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            vals = [c.value for c in row]
            if len(vals) < 5:
                continue
            area, year, code = vals[1], vals[2], vals[3]
            if area != NATION or code != TOTAL_CODE:
                continue
            months = vals[5:17]
            by_year[str(year)] = {MONTH_COLS[i]: (round(float(v), 2) if isinstance(v, (int, float)) else None)
                                   for i, v in enumerate(months)}
        out[sheet] = by_year
    if not out.get("Index"):
        sys.exit("pull_tpso_cpi.py: no national Total rows parsed from the Index sheet — layout changed.")
    return out


def _verify_anchors(sheets):
    bad = []
    for (sheet, year, month), want in ANCHORS.items():
        got = (sheets.get(sheet, {}).get(year, {}) or {}).get(month)
        if got is None or abs(got - want) > 0.005:
            bad.append("%s %s %s: got %s, expected %s" % (sheet, year, month, got, want))
    if bad:
        sys.exit("pull_tpso_cpi.py: ANCHOR FAIL — re-verify before landing:\n  " + "\n  ".join(bad))


def _latest(sheets, sheet):
    by_year = sheets.get(sheet, {})
    for year in sorted(by_year, reverse=True):
        for month in reversed(MONTH_COLS):
            v = by_year[year].get(month)
            if v is not None:
                return year, month, v
    return None, None, None


def build_from_xlsx(xlsx_path, stamp):
    sheets = _parse_xlsx(xlsx_path)
    _verify_anchors(sheets)
    ly_year, ly_month, ly_val = _latest(sheets, "YoY")
    li_year, li_month, li_val = _latest(sheets, "Index")
    return {
        "meta": {
            "title": "Thailand headline CPI / inflation (whole Kingdom, COICOP 2018 basis)",
            "generated_by": "pipeline/pull_tpso_cpi.py",
            "label": "MEASURED — Ministry of Commerce, Trade Policy and Strategy Office (สำนักงาน"
                     "นโยบายและยุทธศาสตร์การค้า, TPSO). NOT a Bank of Thailand series — BOT's own "
                     "real-sector statistics page links out to TPSO for CPI rather than hosting it "
                     "(verified 2026-08-02). Base 2566 (2023) = 100 average.",
            "source": "กระทรวงพาณิชย์ สำนักงานนโยบายและยุทธศาสตร์การค้า (TPSO) — ดัชนีราคาผู้บริโภค/"
                      "เงินเฟ้อ (CPI), COICOP 2018",
            "source_page_url": CPI_PAGE_URL,
            "source_file_url": XLSX_URL,
            "scope": "ประเทศไทย (whole Kingdom), รหัส 0000000 (Total / all items)",
            "sheets": {"Index": "index level, base 2566=100 average",
                       "MoM": "% change vs prior month",
                       "YoY": "% change vs same month prior year — the headline inflation figure",
                       "AoA": "% change, cumulative Jan-to-date average vs same period prior year"},
            "pulled": stamp,
            "latest_yoy": {"year_be": ly_year, "month": ly_month, "value": ly_val, "unit": "% YoY"},
            "latest_index": {"year_be": li_year, "month": li_month, "value": li_val},
            "acceptance_test": "4 spot-verified year/month cells (Index + YoY sheets) checked against "
                               "fixed anchors on every pull; a mismatch hard-fails rather than "
                               "silently landing a different reading — see ANCHORS in "
                               "pull_tpso_cpi.py.",
        },
        "national_total": sheets,
    }


def _dumps(data):
    return json.dumps(data, ensure_ascii=False, indent=2) + "\n"


def main():
    for s in (sys.stdout, sys.stderr):
        try:
            s.reconfigure(encoding="utf-8")
        except Exception:
            pass
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--stamp", default=datetime.date.today().isoformat(),
                    help="YYYY-MM-DD pull date embedded in meta.pulled (default: today)")
    ap.add_argument("--check", action="store_true",
                    help="OFFLINE: re-parse the cached raw XLSX and byte-compare against "
                         "source-data/tpso_cpi.json; exit 1 on drift, exit 3 SKIP if the committed "
                         "JSON or the gitignored raw cache is absent (network-pulled input).")
    args = ap.parse_args()

    if args.check:
        if not os.path.exists(OUT) or not os.path.exists(RAW_XLSX):
            print("pull_tpso_cpi.py --check: SKIP (committed tpso_cpi.json or gitignored raw "
                  "source-data/.tpso_cpi_raw/ absent — network-pulled input, not drift)")
            sys.exit(3)
        prev = json.load(open(OUT, encoding="utf-8"))
        data = build_from_xlsx(RAW_XLSX, prev["meta"]["pulled"])
        if _dumps(data) != open(OUT, encoding="utf-8").read():
            sys.exit("pull_tpso_cpi.py --check: tpso_cpi.json drifted from a fresh parse of the "
                     "cached raw — re-run python3 pipeline/pull_tpso_cpi.py")
        print("pull_tpso_cpi.py --check: OK (byte-exact from cached raw, latest YoY %s %s = %s%%)"
              % (data["meta"]["latest_yoy"]["year_be"], data["meta"]["latest_yoy"]["month"],
                 data["meta"]["latest_yoy"]["value"]))
        return

    os.makedirs(RAW_DIR, exist_ok=True)
    print("fetching %s ..." % XLSX_URL)
    xlsx_bytes = _fetch(XLSX_URL)
    with open(RAW_XLSX, "wb") as f:
        f.write(xlsx_bytes)
    data = build_from_xlsx(RAW_XLSX, args.stamp)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8", newline="\n") as f:
        f.write(_dumps(data))
    print("wrote %s" % OUT)
    ly = data["meta"]["latest_yoy"]
    print("  latest headline CPI YoY: %s %s (BE) = %+.2f%%" % (ly["year_be"], ly["month"], ly["value"]))
    print("  acceptance test: PASSED (4/4 anchor cells matched)")


if __name__ == "__main__":
    main()
