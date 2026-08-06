#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pull_bot_policy_rate.py — MEASURED Bank of Thailand policy rate (owner review 2026-08-02: "doesn't
BOT have this stuff?" — verifying whether the 1% shown on the Macro tab is still current).

WHY NOT THE BOTWEBSTAT REPORT TABLE (reportID=223, FM_RT_001_S2)
  BOT does publish a money-market-rates statistics table that carries the policy rate as its first
  row (BOTWEBSTAT.aspx?reportID=223) — the same idiom pull_bot_uvpi.py / pull_bot_credit.py use for
  their reports. But that specific table's own "valid period range" tops out at July 2025 (verified
  2026-08-02, 13+ months stale) — it is not the live source for this figure. The genuinely current,
  actively-maintained source is BOT's OWN monetary-policy microsite, which republishes a full MPC
  decision history as a downloadable XLSX after every meeting:

    https://www.bot.or.th/th/our-roles/monetary-policy/mpc-publication/policy-interest-rate.html
    -> "ดาวน์โหลดตาราง (ไฟล์ .XLSX)" -> /content/dam/.../policy-interest-rate/table-mpc-<YYYY-BE>-<n>-<hash>.xlsx

  The filename embeds the BE year + meeting number (e.g. table-mpc-2569-3-NVHKxr00.xlsx = the file
  published after MPC meeting 3/2569, 24 June 2026) and changes after every meeting, so this puller
  fetches the HTML page fresh each run to discover the current filename, then downloads that XLSX.
  Sheet1 is every MPC decision back to May 2000 (rate, unanimous/split vote, hold/hike/cut) — the
  policy rate itself is column H ("อัตราดอกเบี้ยนโยบาย\nPolicy rate\n(%)").

VERIFIED 2026-08-02 (live pull): unchanged since MPC meeting 1/2569 (25 Feb 2026, cut 25bps to 1.00%);
held at 2/2569 (29 Apr 2026) and 3/2569 (24 Jun 2026, unanimous 7-0). Next meeting 26 Aug 2026 (not yet
held as of this pull) — so 1.00% is confirmed CURRENT, matching the BIS-mirrored figure already on the
Macro tab (macro_indicators.json policy_rate, period 2026-06).

DETERMINISM: the live pull's XLSX is cached (gitignored) at source-data/.bot_policy_rate_raw/; the
distilled source-data/bot_policy_rate.json is committed. `--check` re-parses the cached raw OFFLINE and
byte-compares. No wall clock in the data — `pulled` comes only from --stamp.

Run:
  python3 pull_bot_policy_rate.py --stamp 2026-08-02   # discover + download + parse + write
  python3 pull_bot_policy_rate.py                       # default --stamp = today
  python3 pull_bot_policy_rate.py --check               # offline byte-reproduce from cached raw
"""
import argparse
import datetime
import json
import os
import re
import sys
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

OUT = os.path.join(ROOT, "source-data", "bot_policy_rate.json")
RAW_DIR = os.path.join(ROOT, "source-data", ".bot_policy_rate_raw")
RAW_XLSX = os.path.join(RAW_DIR, "table-mpc.xlsx")
RAW_SRC_URL = os.path.join(RAW_DIR, "source_url.txt")

PAGE_URL = "https://www.bot.or.th/th/our-roles/monetary-policy/mpc-publication/policy-interest-rate.html"
BOT_BASE = "https://www.bot.or.th"
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"}

# Spot-verification anchors — meetings whose outcome is permanent public record and will not revise.
# A mismatch means the parser broke (column shuffled, sheet renamed), not that history changed —
# stop and re-verify rather than silently landing a different reading.
ANCHORS = {
    "2000-05-23": 1.5,     # ครั้งที่ 1/2543 — first row in the series
    "2001-06-08": 2.5,     # by-Governor 100bps emergency hike
    "2026-02-25": 1.00,    # ครั้งที่ 1/2569 — cut 25bps, the level still in force
    "2026-06-24": 1.00,    # ครั้งที่ 3/2569 — unanimous 7-0 hold, most recent meeting at pull time
}


def _fetch(url, data=None, timeout=60):
    req = urllib.request.Request(url, data=data, headers=UA)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def _discover_xlsx_url(html_bytes):
    html = html_bytes.decode("utf-8", "replace")
    m = re.search(r'href="(/content/dam/bot/documents/th/our-roles/monetary-policy/'
                  r'mpc-publication/policy-interest-rate/table-mpc-[^"]+\.xlsx)"', html)
    if not m:
        sys.exit("pull_bot_policy_rate.py: could not find the table-mpc-*.xlsx download link on "
                 "the policy-interest-rate.html page — page layout changed.")
    return BOT_BASE + m.group(1)


def _parse_xlsx(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        vals = [c.value for c in row]
        d = vals[1]  # column B: meeting date
        if not hasattr(d, "isoformat"):
            continue  # footnote / blank rows carry no date
        rate = vals[7]  # column H: policy rate (%)
        if rate is None:
            continue
        rows.append({
            "meeting": (vals[0] or "").strip() if isinstance(vals[0], str) else None,
            "date": d.date().isoformat(),
            "body": vals[4] or vals[3],           # English body name if present, else Thai
            "decision_th": vals[5], "decision_en": vals[6],
            "rate": round(float(rate), 2),
        })
    rows.sort(key=lambda r: r["date"])
    if not rows:
        sys.exit("pull_bot_policy_rate.py: no dated decision rows parsed from Sheet1 — layout changed.")
    return rows


def _verify_anchors(rows):
    by_date = {r["date"]: r["rate"] for r in rows}
    bad = []
    for date, want in ANCHORS.items():
        got = by_date.get(date)
        if got is None or abs(got - want) > 0.005:
            bad.append("%s: got %s, expected %s" % (date, got, want))
    if bad:
        sys.exit("pull_bot_policy_rate.py: ANCHOR FAIL — re-verify before landing:\n  " + "\n  ".join(bad))


def build_from_xlsx(xlsx_path, stamp, source_url):
    rows = _parse_xlsx(xlsx_path)
    _verify_anchors(rows)
    latest = rows[-1]
    return {
        "meta": {
            "title": "Bank of Thailand policy rate — full MPC decision history",
            "generated_by": "pipeline/pull_bot_policy_rate.py",
            "label": "MEASURED — Bank of Thailand (ธปท.), MPC (กนง.) decision history, republished as "
                     "an XLSX after every meeting on BOT's own monetary-policy microsite. Supersedes "
                     "the BOTWEBSTAT report table (reportID=223, FM_RT_001_S2), which is stuck at "
                     "July 2025 (verified stale 2026-08-02) and not used here.",
            "source": "ธนาคารแห่งประเทศไทย (ธปท.) — ข้อมูลอัตราดอกเบี้ยนโยบายและผลการตัดสินนโยบายการเงิน"
                      "ที่ผ่านมา",
            "source_page_url": PAGE_URL,
            "source_file_url": source_url,
            "definition": "The 1-day bilateral repurchase rate (อัตราดอกเบี้ยธุรกรรมซื้อคืนพันธบัตร"
                          "แบบทวิภาคีระยะ 1 วัน) is BOT's policy rate since 17 Jan 2007; the 14-day "
                          "repo rate was used from 23 May 2000 to 16 Jan 2007.",
            "pulled": stamp,
            "n_meetings": len(rows),
            "period_range": {"min": rows[0]["date"], "max": rows[-1]["date"]},
            "acceptance_test": "4 spot-verified meeting dates (2000-05-23 .. 2026-06-24) checked "
                               "against fixed anchors on every pull; a mismatch hard-fails rather "
                               "than silently landing a different reading — see ANCHORS in "
                               "pull_bot_policy_rate.py.",
        },
        "latest": latest,
        "history": rows,
    }


def _dumps(data):
    return json.dumps(data, ensure_ascii=False, indent=2) + "\n"


def main():
    for s in (sys.stdout, sys.stderr):
        try:
            s.reconfigure(encoding="utf-8")
        except Exception:
            pass
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--stamp", default=datetime.date.today().isoformat(),
                    help="YYYY-MM-DD pull date embedded in meta.pulled (default: today)")
    ap.add_argument("--check", action="store_true",
                    help="OFFLINE: re-parse the cached raw XLSX and byte-compare against "
                         "source-data/bot_policy_rate.json; exit 1 on drift, exit 3 SKIP if the "
                         "committed JSON or the gitignored raw cache is absent (network-pulled input).")
    args = ap.parse_args()

    if args.check:
        if not os.path.exists(OUT) or not os.path.exists(RAW_XLSX):
            print("pull_bot_policy_rate.py --check: SKIP (committed bot_policy_rate.json or "
                  "gitignored raw source-data/.bot_policy_rate_raw/ absent — network-pulled input, "
                  "not drift)")
            sys.exit(3)
        prev = json.load(open(OUT, encoding="utf-8"))
        src_url = open(RAW_SRC_URL, encoding="utf-8").read().strip() if os.path.exists(RAW_SRC_URL) else prev["meta"]["source_file_url"]
        data = build_from_xlsx(RAW_XLSX, prev["meta"]["pulled"], src_url)
        if _dumps(data) != open(OUT, encoding="utf-8").read():
            sys.exit("pull_bot_policy_rate.py --check: bot_policy_rate.json drifted from a fresh "
                     "parse of the cached raw — re-run python3 pipeline/pull_bot_policy_rate.py")
        print("pull_bot_policy_rate.py --check: OK (byte-exact from cached raw, %d meetings)"
              % data["meta"]["n_meetings"])
        return

    os.makedirs(RAW_DIR, exist_ok=True)
    print("fetching %s ..." % PAGE_URL)
    page_html = _fetch(PAGE_URL)
    xlsx_url = _discover_xlsx_url(page_html)
    print("  discovered %s" % xlsx_url)
    xlsx_bytes = _fetch(xlsx_url)
    with open(RAW_XLSX, "wb") as f:
        f.write(xlsx_bytes)
    with open(RAW_SRC_URL, "w", encoding="utf-8", newline="\n") as f:
        f.write(xlsx_url + "\n")
    data = build_from_xlsx(RAW_XLSX, args.stamp, xlsx_url)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8", newline="\n") as f:
        f.write(_dumps(data))
    print("wrote %s" % OUT)
    lt = data["latest"]
    print("  latest meeting %s (%s): %.2f%% — %s" % (lt["date"], lt["meeting"], lt["rate"], lt["decision_en"]))
    print("  acceptance test: PASSED (4/4 anchor meetings matched)")


if __name__ == "__main__":
    main()
