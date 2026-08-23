#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pull_pinksheet.py — standalone monthly World Bank Pink Sheet roll (the missing puller)
=====================================================================================
WHY THIS EXISTS
---------------
The commodity board is the MEASURED global-commodity read that drives the agri portfolio-risk lens
(objective #1). Its source files —

    source-data/commodities.json          five crops   (rice / rubber / palm / sugar / maize)
    source-data/commodities_protein.json  ten protein/forestry/gold series
    source-data/commodity_board.json       the 11-row Overview board (meta.board via derive.py)

— are refreshed off the World Bank CMO "Monthly Prices" xlsx. Until now nothing pulled them on a
schedule: every roll (e.g. the 2026M06→2026M07 refresh on 2026-08-21) was a manual, ~15-step
operation done by hand or noticed opportunistically by the autonomous loop, so the board silently
aged a month between refreshes (energy and fuel prices each already have their own cron; the Pink
Sheet did not). This script is the one-command, self-testing equivalent of that manual roll, so a
monthly workflow (.github/workflows/data-pinksheet.yml) can keep the board current with no human in
the loop and no hand-typed numbers.

WHAT IT DOES (and, as importantly, does NOT)
--------------------------------------------
- Downloads the current CMO-Historical-Data-Monthly.xlsx (scrapes the live WB landing-page link,
  falls back to the last-known hash — the SAME contract autox_enrich_loop.pinksheet_url() uses) and
  parses the "Monthly Prices" sheet with the EXACT same header/row contract as
  autox_enrich_loop.stage_commodities() (hdr=rows[4], data=rows[6:], WB_COMMODITIES header map,
  latest non-null value + 12-month YoY). The map is IMPORTED, never re-typed, so the two parsers
  cannot silently drift apart (same discipline as build_commodity_history.py).
- Rewrites commodities.json / commodities_protein.json for every series the committed files carry
  (a series whose newest value is frozen upstream — e.g. the discontinued Mexican shrimp, last
  2023M10 — is preserved at that frozen point automatically, because the parse takes the latest
  NON-NULL month per column).
- Updates commodity_board.json's MEASURED fields ONLY — each row's `yoy` and `stale` are set from
  its mapped source (via check_commodity_board.BOARD_TO_SOURCE). The editorial `cls` / `note` /
  `seg` / `reg` are LEFT UNTOUCHED: a script must not invent prose or re-colour a row, so those stay
  as the last human set them. `check_commodity_board.py` only guards yoy+stale, so this is exactly
  the surface it protects.
- Bumps the price-vintage token in platform/data/meta.json's `updated` freshness stamp (the first
  YYYY'M'MM token) to the newest crop vintage, leaving the drought date and everything else intact.
- Caches the xlsx to pipeline/.cache/pinksheet.xlsx so the workflow's follow-on
  `build_commodity_history.py --refresh` (history/sparklines) reads the same file.
- Is a NO-OP when nothing moved: it only writes a file whose content actually changed, so a re-run
  in the same vintage leaves the tree byte-identical and the workflow merges nothing.

It does NOT run derive.py / timeseries.py / the downstream fan-out / provenance — those are the
workflow's job (timeseries writes a NEW snapshot, which rederive_drift.py deliberately refuses to do
unattended; the vintage roll is precisely the "someone asked for it" case, so the workflow runs it
explicitly).

USAGE
  python3 pull_pinksheet.py --selftest   # offline: prove the parse + mapping contract, no network
  python3 pull_pinksheet.py              # pull the xlsx and roll the source files (+ meta stamp)
  python3 pull_pinksheet.py --dry-run    # pull + parse, report what WOULD change, write nothing

Exit 0 = success (rolled, or already current). Exit 2 = a mapping/parse invariant broke (loud,
never a silent bad number). Exit 3 = the xlsx could not be fetched or parsed (network/openpyxl) —
SKIP, honest, changes nothing; the workflow reuses the committed snapshot.
"""
import argparse
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
S = os.path.join(ROOT, "source-data")
P = os.path.join(ROOT, "platform", "data")
CACHE_DIR = os.path.join(HERE, ".cache")
XLSX = os.path.join(CACHE_DIR, "pinksheet.xlsx")
SHEET_NAME = "Monthly Prices"

# Reuse the enrichment loop's header->label map + fetch helpers, and the board guard's row->source
# map — NEVER redefine them here, so no third copy of these contracts can drift (same pattern
# build_commodity_history.py and check_commodity_board.py already establish).
from autox_enrich_loop import WB_COMMODITIES, PINKSHEET, PINKSHEET_PAGE, get  # noqa: E402
from check_commodity_board import BOARD_TO_SOURCE  # noqa: E402

# label -> exact "Monthly Prices" header. Every crop/protein label the committed source files carry
# must be here or the run aborts (exit 2) rather than silently dropping a series. WB_COMMODITIES is
# column->label, so we invert it; the three timber/meat series the board never uses but the protein
# file tracks (Malaysian logs, plywood, lamb) are not in WB_COMMODITIES and are added explicitly.
_LABEL_TO_COL = {lab: col for col, lab in WB_COMMODITIES.items()}
_LABEL_TO_COL.update({
    "logs_my": "Logs, Malaysian",
    "plywood": "Plywood",
    "lamb": "Lamb **",
})

CROP_KEYS = ("palm", "maize", "rice", "sugar", "rubber")           # commodities.json order
PROTEIN_KEYS = ("fishmeal", "shrimp", "logs", "logs_my", "sawnwood",
                "plywood", "beef", "chicken", "lamb", "gold")      # commodities_protein.json order


def _numify(x):
    """Exact same numeric guard as autox_enrich_loop.stage_commodities()'s inline lambda."""
    return float(x) if str(x).replace('.', '', 1).replace('-', '').isdigit() else None


def _latest_yoy(months, col_idx, rows):
    """latest non-null value + its month label + 12-month YoY, byte-for-byte the stage_commodities
    formula: li = last non-null, yi = li-12, yoy = 100*(s[li]-s[yi])/s[yi], rounded as committed."""
    s = [_numify(r[col_idx]) for r in rows]
    idxs = [i for i in range(len(s)) if s[i] is not None]
    if not idxs:
        return None
    li = max(idxs)
    yi = li - 12
    yoy = 100 * (s[li] - s[yi]) / s[yi] if yi >= 0 and s[yi] else None
    return {"latest": round(s[li], 3), "date": str(months[li]),
            "yoy": (round(yoy, 1) if yoy is not None else None)}


def parse_sheet(rows):
    """rows = list(ws.iter_rows(values_only=True)) of the Monthly Prices sheet. Returns
    {label: {latest, date, yoy}} for every label in _LABEL_TO_COL, using the committed contract."""
    hdr = rows[4]
    data = rows[6:]
    names = [str(n).strip() if n else "" for n in hdr]
    months = [r[0] for r in data]
    out = {}
    for lab, col in _LABEL_TO_COL.items():
        if col not in names:
            raise ValueError("Pink Sheet header %r (for %r) not found — sheet layout changed" % (col, lab))
        rec = _latest_yoy(months, names.index(col), data)
        if rec is None:
            raise ValueError("no numeric data for %r (%r)" % (lab, col))
        out[lab] = rec
    return out


def _fetch_xlsx():
    """Scrape the current monthly-xlsx link off the WB landing page; fall back to the last-known
    hash. Caches to XLSX. Returns the raw bytes, or None on any network failure."""
    import re as _re
    url = PINKSHEET
    try:
        html = get(PINKSHEET_PAGE, 60).decode("utf-8", "ignore")
        m = _re.findall(r'https?://[^\s"\'<>]*CMO-Historical-Data-Monthly\.xlsx', html)
        if m:
            url = m[0]
    except Exception:
        pass
    try:
        raw = get(url, 120)
        os.makedirs(CACHE_DIR, exist_ok=True)
        with open(XLSX, "wb") as f:
            f.write(raw)
        return raw
    except Exception as e:
        print("pull_pinksheet: could not fetch the CMO xlsx (%s)" % e, file=sys.stderr)
        return None


def _load_rows(raw):
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    return [list(r) for r in wb[SHEET_NAME].iter_rows(values_only=True)]


def _write_json_if_changed(path, obj):
    """Write compact JSON (matching the committed source-file style) only when content differs.
    Returns True when the file changed."""
    new = json.dumps(obj)
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            if f.read() == new:
                return False
    with open(path, "w", encoding="utf-8") as f:
        f.write(new)
    return True


def _roll_board(parsed):
    """Set each board row's yoy+stale from its mapped source; leave cls/note/seg/reg. Returns
    (board_obj, changed_bool). Aborts (exit 2) if a board row maps to a series we did not parse."""
    path = os.path.join(S, "commodity_board.json")
    board = json.load(open(path, encoding="utf-8"))
    changed = False
    for it in board:
        lab = it.get("lab")
        if lab not in BOARD_TO_SOURCE:
            raise ValueError("board row %r has no BOARD_TO_SOURCE mapping" % lab)
        _which, key = BOARD_TO_SOURCE[lab]
        src = parsed.get(key)
        if not src:
            raise ValueError("board row %r maps to %r, absent in the parsed Pink Sheet" % (lab, key))
        if it.get("yoy") != src["yoy"] or it.get("stale") != src["date"]:
            it["yoy"] = src["yoy"]
            it["stale"] = src["date"]
            changed = True
    if changed:
        with open(path, "w", encoding="utf-8") as f:
            f.write(json.dumps(board))
    return board, changed


def _bump_meta_vintage(new_vintage):
    """Replace the first YYYY'M'MM price token in platform/data/meta.json's `updated` stamp with the
    newest crop vintage, preserving the rest (e.g. the drought date). Returns True when changed."""
    path = os.path.join(P, "meta.json")
    with open(path, encoding="utf-8") as f:
        txt = f.read()
    old = json.loads(txt).get("updated", "")
    new = re.sub(r"\d{4}M\d{2}", new_vintage, old, count=1)
    if new == old:
        return False
    # Targeted replace of only the `updated` value in the raw file text, so every other byte of
    # meta.json (a large derive.py-owned sidecar written ensure_ascii=False) is preserved exactly —
    # never reformat the whole file here. derive.py re-canonicalises meta.json downstream regardless,
    # but this keeps the puller's own write a clean one-token edit.
    old_json = json.dumps(old, ensure_ascii=False)
    new_json = json.dumps(new, ensure_ascii=False)
    # The stamp echoes `updated` a second time (derive.py writes it top-level AND in meta.meta), so
    # the value legitimately appears more than once; bump every copy so they stay in sync. The full
    # dated string is distinctive enough that no unrelated field matches it.
    if old_json not in txt:
        raise ValueError("meta.json `updated` value not found for an in-place bump")
    with open(path, "w", encoding="utf-8") as f:
        f.write(txt.replace(old_json, new_json))
    return True


def selftest():
    """Offline proof of the parse + mapping contract — no network. Builds a tiny synthetic sheet in
    the committed layout, runs the real parser, and asserts the invariants that keep a bad number
    from ever shipping."""
    # every committed source-file key must have a column mapping
    for k in CROP_KEYS + PROTEIN_KEYS:
        assert k in _LABEL_TO_COL, "unmapped series %r — would be silently dropped" % k
    # every board row must map to a series we parse
    for lab, (_w, key) in BOARD_TO_SOURCE.items():
        assert key in _LABEL_TO_COL, "board %r -> %r is not a parseable column" % (lab, key)
    # synthetic Monthly Prices sheet: 4 filler rows, header at idx 4, blank unit row 5, data from 6.
    cols = ["Commodity"] + [_LABEL_TO_COL[k] for k in CROP_KEYS + PROTEIN_KEYS]
    rows = [["x"] * len(cols) for _ in range(4)]
    rows.append(cols)                       # hdr = rows[4]
    rows.append(["($)"] * len(cols))        # rows[5] skipped
    # 14 months so a 12-month YoY exists; value = 100 + month for the first crop, flat 50 elsewhere
    for mo in range(1, 15):
        r = ["2025M%02d" % mo if mo <= 12 else "2026M%02d" % (mo - 12)]
        for j, k in enumerate(CROP_KEYS + PROTEIN_KEYS):
            r.append(100 + mo if j == 0 else 50)
        rows.append(r)
    parsed = parse_sheet(rows)
    assert set(parsed) == set(CROP_KEYS + PROTEIN_KEYS)
    first = parsed[CROP_KEYS[0]]
    assert first["date"] == "2026M02", first          # newest month label
    assert first["latest"] == 114.0, first            # 100 + 14
    # YoY vs 12 months earlier (row 2, value 102): 100*(114-102)/102 = 11.8
    assert first["yoy"] == 11.8, first
    flat = parsed[CROP_KEYS[1]]
    assert flat["yoy"] == 0.0, flat                   # constant series -> 0% YoY
    print("selftest OK: parse contract + %d series mapped, board rows all resolvable"
          % len(_LABEL_TO_COL))
    return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--selftest", action="store_true", help="offline parse/mapping proof, no network")
    ap.add_argument("--dry-run", action="store_true", help="pull + parse, report changes, write nothing")
    args = ap.parse_args()

    if args.selftest:
        return selftest()

    raw = _fetch_xlsx()
    if raw is None:
        print("pull_pinksheet: SKIP (xlsx unfetchable) — committed board reused, nothing written")
        return 3
    try:
        parsed = parse_sheet(_load_rows(raw))
    except Exception as e:
        print("pull_pinksheet: SKIP (parse failed: %s) — nothing written" % e, file=sys.stderr)
        return 3

    crops = {k: parsed[k] for k in CROP_KEYS}
    protein = {k: parsed[k] for k in PROTEIN_KEYS}
    new_vintage = max(crops[k]["date"] for k in CROP_KEYS)   # crops share the newest month

    if args.dry_run:
        cur = json.load(open(os.path.join(S, "commodities.json"), encoding="utf-8"))
        moved = [k for k in CROP_KEYS if cur.get(k) != crops[k]]
        print("pull_pinksheet --dry-run: newest crop vintage %s; crops that would move: %s"
              % (new_vintage, ", ".join(moved) or "none"))
        return 0

    ch1 = _write_json_if_changed(os.path.join(S, "commodities.json"), crops)
    ch2 = _write_json_if_changed(os.path.join(S, "commodities_protein.json"), protein)
    _board, ch3 = _roll_board(parsed)
    ch4 = _bump_meta_vintage(new_vintage)

    if not (ch1 or ch2 or ch3 or ch4):
        print("pull_pinksheet: already current at %s — nothing changed" % new_vintage)
        return 0
    print("pull_pinksheet: rolled Pink Sheet to %s "
          "(commodities=%s protein=%s board=%s meta=%s). "
          "Cached xlsx -> %s for build_commodity_history --refresh."
          % (new_vintage, ch1, ch2, ch3, ch4, os.path.relpath(XLSX, ROOT)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
