#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Project the owner-side underwriting Excel into the MINIMAL policy reference the digest needs.

WHY THIS EXISTS
---------------
The rival digest reported what competitors did and stopped there. Every line was true and none
of them named a thing we could change, because the email had no idea what our own policy allows.
An "action" is only actionable against a lever, and our levers live in the underwriting pack —
outside this repo, on purpose, as restricted internal policy material.

This script is the bridge, and it follows the `ingest_real_tape.py` pattern exactly: the source
workbook stays OFF-REPO and owner-side, and only a small derived projection is committed.

WHAT IT DELIBERATELY DOES NOT CARRY
-----------------------------------
The workbook holds far more than this: the approval-authority ladder, escalation routing, the
fraud ต้นขั้ว thresholds, income-margin tables, the income-doc checklist by lane, NCB rules.
NONE of it is extracted. A competitive comparison needs the collateral-side ceilings and nothing
else, so that is all this reads. Keeping the extracted surface to the minimum is the point, not
an oversight — if a future comparison genuinely needs another sheet, add it deliberately and say
why here.

The output lands in `pipeline/`, NOT in `platform/data/`. `platform/` is the deployed folder;
anything placed there is served. `pipeline/` is not part of the deploy (repo-root vercel.json sets
outputDirectory=platform), and this repo is private, so the projection stays inside the same trust
boundary as the rest of the owner-side material already here.

USAGE
-----
    python pipeline/build_policy_levers.py --src "<path to AutoX_LTV_Database_COMPLETE.xlsx>"
    AUTOX_LTV_XLSX=<path> python pipeline/build_policy_levers.py

Like the tape ingest, this is NOT in the determinism gate — its input is off-repo, so CI cannot
reproduce it. Everything downstream of the committed JSON is gated.
"""
from __future__ import annotations

import argparse
import io
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "policy_levers.json")

# Sheets we read. Anything not on this list is never opened — see the module docstring.
SHEET_LTV = "Programs_Base"
SHEET_ELIG = "Collateral_Eligibility"
SHEET_PROMO = "Programs_Promo"      # the live LTV-uplift campaign (ประกาศ 047 / 107, ext. ปกค 125)

# Thai Buddhist-era years land in the workbook as bare two- or four-digit BE. "14 ม.ค. 70" is
# 2570 BE = 2027 CE; left alone it reads as either 1970 or 542 years in the future, and a campaign
# window is exactly the field where that matters.
TH_MON = {"ม.ค.": 1, "ก.พ.": 2, "มี.ค.": 3, "เม.ย.": 4, "พ.ค.": 5, "มิ.ย.": 6,
          "ก.ค.": 7, "ส.ค.": 8, "ก.ย.": 9, "ต.ค.": 10, "พ.ย.": 11, "ธ.ค.": 12}

# The `sub` column packs the lane's own caps into one Thai string, e.g.
# "วงเงิน ≤700,000 · 24%/ปี · ถือครอง 90 วัน". Pulling the three numbers out beats re-typing
# them, because a re-typed constant is a constant that drifts from the workbook silently.
RE_CAP = re.compile(r"วงเงิน\s*[≤<]=?\s*([\d,]+)")
RE_RATE = re.compile(r"([\d.]+)\s*%\s*/\s*ปี")
RE_HOLD = re.compile(r"ถือครอง\s*(\d+)\s*วัน")


def _num(s):
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    m = re.search(r"([\d.]+)", str(s).replace(",", ""))
    return float(m.group(1)) if m else None


def _pct(s):
    """'75%' / 75 / 0.75 → 75.0. The workbook mixes all three spellings across sheets."""
    v = _num(s)
    if v is None:
        return None
    return round(v * 100, 2) if v <= 1.0 else round(v, 2)


def _age_band(s):
    """'21–25 ปี' → (21, 25); '≤20 ปี' → (1, 20). None when the cell is not a band.

    The single-number spelling is an UPPER BOUND, not a one-year band: the workbook writes the
    widest car lane as '≤20 ปี', meaning one year old through twenty. Reading it as (20, 20)
    turned that lane into "cars aged exactly 20", which is both wrong and the kind of wrong that
    reads as a real policy — so the ≤ / <= / ไม่เกิน forms are handled explicitly.
    """
    if not s:
        return None
    t = str(s).replace("–", "-").replace("—", "-")
    m = re.search(r"(\d+)\s*-\s*(\d+)", t)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"(\d+)", t)
    if not m:
        return None
    n = int(m.group(1))
    if re.search(r"[≤<]|ไม่เกิน", t):
        return 1, n
    if re.search(r"[≥>]|ตั้งแต่", t):
        return n, None
    return n, n


def read_ltvx(path):
    """Parse the LTVX / C-code table out of the Product Selector, the stated source of truth.

    WHAT LTVX ACTUALLY IS — worth stating, because the name misleads. It is not a rate promotion
    and not a sheet in the workbook. It is a BOOKING MODE: every brand in the Redbook has a
    duplicate item prefixed "LTVX", and choosing the twin is what lets Mobius write an LTV above
    100%. The plain item locks at 100%. The twin is only legal when the case resolves to a C-code,
    so the booking mode and the credit programme are one decision.

    A FINDING IN ITS OWN RIGHT: this table lives ONLY in the Selector's JavaScript. It is not in
    AutoX_LTV_Database_COMPLETE.xlsx, so it sits outside the Excel-is-the-record-of-truth contract
    that governs every other policy number — exactly the double-maintenance risk that the pack's
    own EXCEL-ARCHITECTURE.md warns about. Parsing it here rather than re-typing it means at least
    this consumer cannot drift from the Selector; it does not fix the underlying gap.

    Returns [] when the selector is not supplied, so the digest degrades to the base grid.
    """
    if not path or not os.path.exists(path):
        return []
    with io.open(path, encoding="utf-8", errors="replace") as f:
        s = f.read()
    pat = re.compile(
        r"code:\s*(?:'([A-Z]{4})'|isTitle\?'([A-Z]{4})':'([A-Z]{4})')\s*,\s*"
        r"cap:\s*(\d+)\s*,\s*prog:\s*'([A-Z]+)'\s*,\s*rate:\s*([\d.]+)")
    out, seen = [], set()
    for m in pat.finditer(s):
        lit, t_code, h_code, cap, prog, rate = m.groups()
        for code, kind in ([(lit, "hp_bank")] if lit else
                           [(t_code, "title"), (h_code, "hp_nonbank")]):
            if not code or code in seen:
                continue
            seen.add(code)
            out.append({"code": code, "loan_kind": kind, "cap_pct": int(cap),
                        "rate_pct": float(rate), "risk": prog})
    # The low-risk gate, quoted rather than paraphrased — an action that names a gate has to name
    # it correctly, and this one changed in v46.
    gate = None
    g = re.search(r"lowGates\s*=\s*([^;]{0,160});", s)
    if g:
        gate = g.group(1).strip()
    vtypes = re.findall(r"p\.vtype===?'([A-Z]+)'", s)
    return sorted(out, key=lambda r: (-r["cap_pct"], r["rate_pct"])), gate, sorted(set(vtypes))


# The Selector's own identifiers, spelled out. The verbatim expression is kept as the record —
# a gate should be quoted, not paraphrased — but `gAA&&gVeh&&occ&&age&&h12` reached the digest
# and would have reached the owner's inbox, where it means nothing. Both forms ship: the
# expression for whoever checks it against the Selector, the sentence for whoever reads the email.
GATE_TERMS = [
    ("gAA", "NCB grade AA"),
    ("gVeh", "an eligible vehicle type (PA/PU/VAN)"),
    ("occ", "salaried occupation"),
    ("age", "borrower aged 31–60"),
    ("h12", "at least 12 months of credit history"),
]


def gate_readable(expr):
    """'gAA&&gVeh&&occ&&age&&h12' → a sentence. Unknown identifiers are surfaced, not dropped:
    if v47 adds a term this must not quietly report the old gate as though it still held."""
    if not expr:
        return None
    found = [txt for ident, txt in GATE_TERMS if re.search(r"\b%s\b" % re.escape(ident), expr)]
    known = {ident for ident, _ in GATE_TERMS}
    extra = [t for t in re.findall(r"[A-Za-z_][A-Za-z0-9_]*", expr) if t not in known]
    if not found:
        return None
    out = " + ".join(found)
    if extra:
        out += " + %s (unnamed in this build — check the Selector)" % ", ".join(sorted(set(extra)))
    return out


def _be_date(s):
    """'สิ้นสุด 14 ม.ค. 70' → '2027-01-14'. None when the cell carries no dated window."""
    if not s:
        return None
    t = str(s)
    m = re.search(r"(\d{1,2})\s*(" + "|".join(re.escape(k) for k in TH_MON) + r")\s*(\d{2,4})", t)
    if not m:
        return None
    d, mon, yr = int(m.group(1)), TH_MON[m.group(2)], int(m.group(3))
    be = yr + 2500 if yr < 100 else yr            # 70 → 2570
    return "%04d-%02d-%02d" % (be - 543, mon, d)  # BE → CE


def read(src):
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required: pip install --break-system-packages openpyxl")
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    for need in (SHEET_LTV, SHEET_ELIG):
        if need not in wb.sheetnames:
            sys.exit("workbook has no sheet %r — is this the right file?" % need)

    def rows(name):
        ws = wb[name]
        it = ws.iter_rows(values_only=True)
        head = [str(h).strip() if h is not None else "" for h in next(it)]
        for r in it:
            rec = {head[i]: r[i] for i in range(min(len(head), len(r)))}
            if any(v is not None and str(v).strip() for v in rec.values()):
                yield rec

    # ---- LTV ceilings, one record per (program, collateral, tier, brand, age band) -----------
    ltv, lanes = [], {}
    for r in rows(SHEET_LTV):
        pct = _pct(r.get("ltv"))
        if pct is None:
            continue
        prog = (r.get("program") or "").strip()
        sub = str(r.get("sub") or "")
        if prog and prog not in lanes:
            cap, rate, hold = RE_CAP.search(sub), RE_RATE.search(sub), RE_HOLD.search(sub)
            lanes[prog] = {
                "program": prog,
                "collateral": (r.get("vehicle") or "").strip(),
                "model": (r.get("model") or "").strip(),
                "loan_cap_thb": int(cap.group(1).replace(",", "")) if cap else None,
                "rate_pct_yr": float(rate.group(1)) if rate else None,
                "holding_days": int(hold.group(1)) if hold else None,
                "basis": (r.get("badge") or "").strip(),
                "note": (r.get("note") or "").strip() or None,
            }
        band = _age_band(r.get("vehicle_age"))
        ltv.append({
            "program": prog,
            "collateral": (r.get("vehicle") or "").strip(),
            "tier": (r.get("tier") or "").strip(),
            "brands": (r.get("brand") or "").strip(),
            "age_from": band[0] if band else None,
            "age_to": band[1] if band else None,
            "ltv_pct": pct,
        })

    # ---- Eligibility: the gate BEFORE any LTV applies ----------------------------------------
    elig = []
    for r in rows(SHEET_ELIG):
        age = _num(r.get("max_age_years"))
        days = _num(r.get("ownership_days"))
        elig.append({
            "collateral": (str(r.get("collateral") or "")).strip(),
            "tier": (str(r.get("tier") or "")).strip(),
            "max_age_years": int(age) if age else None,
            "ownership_days": int(days) if days else None,
            "src": (str(r.get("src") or "")).strip(),
        })

    # ---- THE LIVE CAMPAIGN — the LTV uplift on top of the base grid ---------------------------
    # This sheet is why a competitive LTV comparison is possible at all: the base grid is what the
    # ระเบียบ allows, the campaign is what we are actually offering right now. Note it carries
    # THREE ltv columns, not one — ปกค 125 split the lanes by whether insurance is taken
    # (ltv_no / ltv_yes), with ltv_merged for the rows that predate the split. Reading only `ltv`
    # (which does not exist here) silently returns an empty campaign, which would have read as
    # "we run no campaign" — the exact opposite of the truth.
    promo, windows = [], {}
    if SHEET_PROMO in wb.sheetnames:
        for r in rows(SHEET_PROMO):
            cand = [(_pct(r.get(k)), k) for k in ("ltv_merged", "ltv_yes", "ltv_no")]
            cand = [(v, k) for v, k in cand if v is not None]
            if not cand:
                continue
            best = max(cand)
            band = _age_band(r.get("vehicle_age"))
            period = str(r.get("period") or "").strip()
            ends = _be_date(period)
            if ends:
                windows[ends] = windows.get(ends, 0) + 1
            promo.append({
                "program": (r.get("program") or "").strip(),
                "collateral": (r.get("vehicle") or "").strip(),
                "model": (r.get("model") or "").strip(),
                "tier": (r.get("tier") or "").strip(),
                "brands": (r.get("brand") or "").strip(),
                "age_from": band[0] if band else None,
                "age_to": band[1] if band else None,
                "ltv_pct": best[0],
                "ltv_basis": {"ltv_yes": "with insurance", "ltv_no": "without insurance",
                              "ltv_merged": "insurance-agnostic"}[best[1]],
                "period_th": period or None,
                "ends": ends,
                "ref": str(r.get("ref") or "").strip() or None,
            })

    return ltv, elig, sorted(lanes.values(), key=lambda x: x["program"]), promo, windows


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--src", default=os.environ.get("AUTOX_LTV_XLSX"),
                    help="path to AutoX_LTV_Database_COMPLETE.xlsx (owner-side, off-repo)")
    ap.add_argument("--selector", default=os.environ.get("AUTOX_SELECTOR_HTML"),
                    help="AutoX_Product_Selector_Wizard_*.html — the source of truth for the "
                         "LTVX / C-code table, which is NOT in the workbook")
    ap.add_argument("--out", default=OUT)
    a = ap.parse_args()
    if not a.src:
        sys.exit("need --src <xlsx> or AUTOX_LTV_XLSX — the workbook is owner-side, not in this repo")
    if not os.path.exists(a.src):
        sys.exit("no such file: %s" % a.src)

    ltv, elig, lanes, promo, windows = read(a.src)
    if not ltv:
        sys.exit("read 0 LTV rows — refusing to write an empty projection over a good one")

    ceilings = [r["ltv_pct"] for r in ltv]

    # ---- LTVX / C-code: the booking mode that lifts the Mobius cap above 100% -----------------
    codes, gate, vtypes = read_ltvx(a.selector) if a.selector else ([], None, [])
    title = [c for c in codes if c["loan_kind"] == "title"]
    ltvx = {
        "what": "A booking mode, not a rate promotion. Every Redbook brand has a duplicate item "
                "prefixed LTVX; choosing the twin is what lets Mobius write LTV above 100%. The "
                "plain item locks at 100%. The twin is legal only when the case resolves to a "
                "C-code, so booking mode and credit programme are one decision.",
        "source": os.path.basename(a.selector) if a.selector else None,
        "source_note": "Parsed from the Product Selector's JavaScript. This table is NOT in the "
                       "workbook, so it sits outside the Excel-is-the-record-of-truth contract "
                       "that governs every other policy number here.",
        "standard_cap_pct": 100,
        "codes": codes,
        "low_risk_gate": gate,
        "low_risk_gate_readable": gate_readable(gate),
        "eligible_vehicle_types": vtypes,
        "v46_change": "v46 (Aug 2026) moved the low-risk collateral test from BRAND TIER to "
                      "VEHICLE TYPE. Low-risk previously needed a MAJOR-tier brand (4 brands in "
                      "the ratebook); now every PA / PU / VAN qualifies regardless of tier. This "
                      "is the change that widened who can be booked LTVX.",
        # What a title-loan case can actually reach — our core book.
        "title_cap_pct": max([c["cap_pct"] for c in title], default=None),
        "title_best_rate_pct": min([c["rate_pct"] for c in title], default=None),
        "title_rate_at_cap": next((c["rate_pct"] for c in
                                   sorted(title, key=lambda x: (-x["cap_pct"], x["rate_pct"]))),
                                  None),
    }
    # The campaign ceiling is what a competitor comparison must use — a rival advertising its best
    # LTV is advertising its campaign, so netting it against our BASE grid would understate us.
    dated = [p for p in promo if p.get("ends")]
    camp = {
        "n_rows": len(promo),
        "n_programs": len(set(p["program"] for p in promo if p["program"])),
        "ltv_max_pct": max([p["ltv_pct"] for p in promo], default=None),
        # The expiry that matters is the CEILING lane's own window, not the most common one across
        # the sheet. The 100% RHUC pickup lanes end 01 ต.ค. 69 while the 80% lanes run to
        # 14 ม.ค. 70 — quoting the modal date would attach the wrong expiry to the headline number,
        # and this specific number is the one with a hard stop that is known not to extend.
        "ceiling_ends": None,
        "windows": dict(sorted(windows.items())),
        "refs": sorted(set(p["ref"] for p in dated if p["ref"])),
        "uplift_pp": None,
        "top": [],
    }
    if camp["ltv_max_pct"] is not None:
        camp["uplift_pp"] = round(camp["ltv_max_pct"] - max(ceilings), 2)
        top = max(p["ltv_pct"] for p in promo)
        ends_at_top = sorted(set(p["ends"] for p in promo
                                 if p["ltv_pct"] >= top and p.get("ends")))
        camp["ceiling_ends"] = ends_at_top[0] if ends_at_top else None
        camp["ceiling_refs"] = sorted(set(p["ref"] for p in promo
                                          if p["ltv_pct"] >= top and p.get("ref")))
        seen = set()
        for p in sorted(promo, key=lambda x: (-x["ltv_pct"], x["model"])):
            if p["ltv_pct"] < top:
                break
            k = (p["collateral"], p["model"])
            if k in seen:
                continue
            seen.add(k)
            camp["top"].append({"collateral": p["collateral"], "model": p["model"],
                                "tier": p["tier"], "age_from": p["age_from"],
                                "age_to": p["age_to"], "ltv_pct": p["ltv_pct"],
                                "ltv_basis": p["ltv_basis"], "ref": p["ref"]})
    doc = {
        "meta": {
            "title": "AutoX underwriting levers — the collateral-side ceilings only",
            "generated_by": "pipeline/build_policy_levers.py",
            "source": "AutoX_LTV_Database_COMPLETE.xlsx (owner-side; the record of truth per "
                      "production/EXCEL-ARCHITECTURE.md)",
            "source_basename": os.path.basename(a.src),
            "provenance": "MEASURED — read directly from the policy workbook, not inferred.",
            "scope_note": "Collateral ceilings + eligibility ONLY. The approval ladder, fraud "
                          "thresholds, income margins, NCB rules and the income-doc checklist are "
                          "deliberately not extracted.",
            "basis_note": "LTV here is AutoX's own basis (a percentage of the appraised collateral "
                          "value under ระเบียบ ver 9.0). A rival's advertised LTV may be quoted on a "
                          "different base entirely, so the two are NOT comparable until that rival's "
                          "basis is confirmed. Never net one against the other without it.",
            "n_ltv_rows": len(ltv),
            "n_eligibility_rows": len(elig),
            "ltv_min_pct": min(ceilings),
            "ltv_max_pct": max(ceilings),
            "campaign_note": "`campaign` is the LIVE uplift over the base grid. Use its ceiling, "
                             "not the base one, for any competitor comparison — a rival's headline "
                             "LTV is its campaign too.",
            "not_served": "Lives in pipeline/, never platform/ — platform/ is the deployed folder.",
        },
        "ltvx": ltvx,
        "campaign": camp,
        # DELIBERATELY NOT EMITTED: the 103 raw Programs_Promo rows. They were 27.7KB of the 53KB
        # file — over half of it — and grep found no consumer anywhere in the repo. A committed
        # verbatim copy of an internal promo sheet (per-brand, per-model, per-ประกาศ) that nothing
        # reads is pure restricted-material surface area: it would age silently against the
        # workbook, and the whole point of the scope_note above is that only what a consumer needs
        # gets extracted. `campaign` and `lanes` are the derived summaries the digest actually
        # reads; `promo` stays in memory to compute them and is dropped here.
        "lanes": lanes,
        "ltv_ceilings": ltv,
        "eligibility": elig,
    }

    with io.open(a.out, "w", encoding="utf-8", newline="\n") as f:
        json.dump(doc, f, ensure_ascii=False, indent=1, sort_keys=True)
        f.write("\n")
    print("wrote %s — %d LTV rows (%.0f–%.0f%%), %d eligibility rows, %d lanes"
          % (a.out, len(ltv), min(ceilings), max(ceilings), len(elig), len(lanes)))


if __name__ == "__main__":
    main()
