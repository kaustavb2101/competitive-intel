#!/usr/bin/env python3
"""
pull_oae_prices.py — WEEKLY OAE FARM-GATE PRICE pull (objective #1, plan item P3).

Pulls fresh Thai farm-gate prices ("ราคาที่เกษตรกรขายได้" — prices farmers actually
received) from the OAE CKAN catalog (catalog.oae.go.th) and lands them as
source-data/oae_farmgate_prices.json in the schema build_crop_stress.py prefers
over the World Bank GLOBAL proxy. This replaces the stale BE-2562 (2019 CE)
crop_prices.json vintage with a CURRENT measured series — the gap was freshness,
not existence (docs/DATA_ACQUISITION_PLAN.md §2).

FLOW (network — run from a host that can reach catalog.oae.go.th; GitHub Actions
runners can, some sandboxes cannot):
  1. package_search?q=ราคาที่เกษตรกรขายได้ — find candidate datasets (NEVER a
     hardcoded package id; the plan forbids guessing one).
  2. Pick the newest matching package (metadata_modified desc), package_show it.
  3. Download its best resource (CSV > XLS/XLSX > CKAN datastore fallback).
  4. Parse into per-crop monthly series (long OR wide/month-column layouts,
     BE→CE year conversion, Thai month names, comma-grouped numbers, utf-8-sig /
     cp874 decoding), match series labels to the 6 target crops
     (rice, rubber, sugarcane, oil palm, cassava, maize) by Thai keyword priority.
  5. Compute YoY from the series ITSELF (latest period vs same period prior year)
     and write the landing file with full provenance meta. The label is MEASURED:
     these are real farm-gate prices, nothing modelled.

DETERMINISM / PROVENANCE RULES:
  - `pulled` in the meta comes ONLY from --stamp (required for a real write);
    no wall clock is ever embedded in data values, so a re-run with the same
    upstream data and the same --stamp is byte-identical.
  - vintage is read FROM THE DATA (latest BE/CE period present), never assumed.
  - If fewer than MIN_CROPS target crops get a usable YoY the pull FAILS LOUDLY
    and writes nothing — a junk landing file must never demote the honest proxy.

OFFLINE MODES (work with zero egress):
  --selftest   parse two embedded sample fixtures (long CSV + wide month-column
               CSV) through the SAME parse path as the live pull and validate the
               output schema + exact expected YoY numbers. Proves the parse logic.
  --dry-run    (network) search + list resources only; no download parse, no write.

Run:
  python3 pull_oae_prices.py --stamp 2026-07-06            # real pull + write
  python3 pull_oae_prices.py --dry-run                     # search + list only
  python3 pull_oae_prices.py --selftest                    # offline parse proof
"""
import argparse
import csv
import io
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEFAULT_OUT = os.path.join(ROOT, "source-data", "oae_farmgate_prices.json")

BASE = "https://catalog.oae.go.th"
SEARCH_TERM = "ราคาที่เกษตรกรขายได้"
UA = {"User-Agent": "autox-credit-intel/1.0 (data pipeline; contact: repo owner)"}
TIMEOUT = 90
RETRIES = 3
BACKOFF = 2.0          # seconds; doubles each retry
MIN_CROPS = 2          # refuse to land a file with fewer usable target crops
MAX_DATASTORE_ROWS = 32000

# Target crops -> Thai series-label keywords, in PRIORITY order (first pattern
# that matches any label wins; the canonical benchmark series is listed first,
# broader fallbacks after). Matching is substring, whitespace-normalized.
CROP_PATTERNS = {
    "rice":      ["ข้าวเปลือกเจ้าความชื้น 15", "ข้าวเปลือกเจ้าความชื้น15",
                  "ข้าวเปลือกเจ้า", "ข้าวเปลือกหอมมะลิ", "ข้าวเปลือก"],
    "rubber":    ["ยางแผ่นดิบชั้น 3", "ยางแผ่นดิบชั้น3", "ยางแผ่นดิบ",
                  "น้ำยางสด", "ยางพารา"],
    "sugarcane": ["อ้อยโรงงาน", "อ้อย"],
    "oilpalm":   ["ผลปาล์มน้ำมันทั้งทะลาย", "ผลปาล์มน้ำมัน", "ปาล์มน้ำมัน"],
    "cassava":   ["หัวมันสำปะหลังสดคละ", "หัวมันสำปะหลังสด", "หัวมันสำปะหลัง",
                  "มันสำปะหลัง"],
    "maize":     ["ข้าวโพดเลี้ยงสัตว์ความชื้น 14.5", "ข้าวโพดเลี้ยงสัตว์",
                  "ข้าวโพด"],
}
CROP_EN = {"rice": "Rice", "rubber": "Rubber", "sugarcane": "Sugarcane",
           "oilpalm": "Oil palm", "cassava": "Cassava", "maize": "Maize"}

# Thai month names (full + common abbreviations) -> month number
THAI_MONTHS = {
    "มกราคม": 1, "ม.ค.": 1, "ม.ค": 1,
    "กุมภาพันธ์": 2, "ก.พ.": 2, "ก.พ": 2,
    "มีนาคม": 3, "มี.ค.": 3, "มี.ค": 3,
    "เมษายน": 4, "เม.ย.": 4, "เม.ย": 4,
    "พฤษภาคม": 5, "พ.ค.": 5, "พ.ค": 5,
    "มิถุนายน": 6, "มิ.ย.": 6, "มิ.ย": 6,
    "กรกฎาคม": 7, "ก.ค.": 7, "ก.ค": 7,
    "สิงหาคม": 8, "ส.ค.": 8, "ส.ค": 8,
    "กันยายน": 9, "ก.ย.": 9, "ก.ย": 9,
    "ตุลาคม": 10, "ต.ค.": 10, "ต.ค": 10,
    "พฤศจิกายน": 11, "พ.ย.": 11, "พ.ย": 11,
    "ธันวาคม": 12, "ธ.ค.": 12, "ธ.ค": 12,
}

# header keyword sets for column detection (lowercased substring match)
H_LABEL = ("ชนิดสินค้า", "สินค้า", "ชนิด", "รายการ", "commodity", "product")
H_YEAR = ("ปี", "year")
H_MONTH = ("เดือน", "month")
H_DATE = ("วันที่", "งวด", "date", "period")
H_PRICE = ("ราคา", "price", "value", "มูลค่า")
H_UNIT = ("หน่วย", "unit")


class PullError(RuntimeError):
    """A clear, user-facing failure of the pull (message tells you what to fix)."""


# ── HTTP with retries/backoff ────────────────────────────────────────────────
def http_get(url, timeout=TIMEOUT):
    last = None
    for attempt in range(RETRIES):
        try:
            req = urllib.request.Request(url, headers=UA)
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read()
        except Exception as e:  # URLError / HTTPError / timeout / TLS
            last = e
            wait = BACKOFF * (2 ** attempt)
            print("  retry %d/%d after error: %s (sleep %.0fs)"
                  % (attempt + 1, RETRIES, e, wait), file=sys.stderr)
            time.sleep(wait)
    raise PullError("GET failed after %d attempts: %s\n  last error: %s"
                    % (RETRIES, url, last))


def ckan(action, **params):
    url = "%s/api/3/action/%s?%s" % (BASE, action, urllib.parse.urlencode(params))
    raw = http_get(url)
    try:
        doc = json.loads(raw.decode("utf-8"))
    except Exception as e:
        raise PullError("CKAN %s returned non-JSON (%s). First bytes: %r"
                        % (action, e, raw[:200]))
    if not doc.get("success"):
        raise PullError("CKAN %s returned success=false: %s"
                        % (action, json.dumps(doc.get("error"), ensure_ascii=False)[:400]))
    return doc["result"]


# ── dataset / resource selection ─────────────────────────────────────────────
def find_package():
    """package_search for the farm-gate term; newest matching dataset wins."""
    res = ckan("package_search", q=SEARCH_TERM, rows=20)
    results = res.get("results") or []
    if not results:
        raise PullError("package_search found ZERO datasets for %r on %s — "
                        "the catalog may have moved; fall back to the data.go.th "
                        "crop_price_oae topic from a Thai IP (autox_dgt_ingest.py)."
                        % (SEARCH_TERM, BASE))

    def is_match(p):
        hay = " ".join([p.get("title") or "", p.get("name") or "",
                        p.get("notes") or ""])
        return (SEARCH_TERM in hay) or ("ราคา" in hay and "เกษตรกร" in hay)

    matches = [p for p in results if is_match(p)] or results
    matches.sort(key=lambda p: p.get("metadata_modified") or "", reverse=True)
    pkg = matches[0]
    print("dataset: %s  (id=%s, modified=%s)"
          % (pkg.get("title"), pkg.get("id"), pkg.get("metadata_modified")))
    return ckan("package_show", id=pkg["id"])


def rank_resource(r):
    fmt = (r.get("format") or "").strip().lower()
    url = (r.get("url") or "").lower()
    if fmt == "csv" or url.endswith(".csv"):
        return 0
    if fmt in ("xlsx", "xls") or url.endswith((".xlsx", ".xls")):
        return 1
    if r.get("datastore_active"):
        return 2
    return 9


def pick_resource(pkg):
    resources = pkg.get("resources") or []
    if not resources:
        raise PullError("dataset %r has no resources at all" % pkg.get("title"))
    ranked = sorted(resources, key=lambda r: (rank_resource(r),
                                              -(len(r.get("last_modified") or ""))))
    best = ranked[0]
    if rank_resource(best) == 9:
        raise PullError("dataset %r has no CSV/XLSX/datastore resource; formats seen: %s"
                        % (pkg.get("title"),
                           [r.get("format") for r in resources]))
    return best


# ── decoding / row extraction ────────────────────────────────────────────────
def decode_text(raw):
    """UTF-8 (with BOM) first; Thai legacy cp874/tis-620 fallback."""
    for enc in ("utf-8-sig", "cp874", "tis-620"):
        try:
            return raw.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return raw.decode("utf-8", errors="replace")


def rows_from_csv(raw):
    text = decode_text(raw)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    return [[("" if c is None else str(c)).strip() for c in r] for r in rows if any(r)]


def rows_from_xlsx(raw):
    try:
        import openpyxl
    except ImportError:
        raise PullError("resource is XLSX but openpyxl is not installed — "
                        "pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if any(cells):
            rows.append(cells)
    return rows


def rows_from_datastore(resource_id):
    res = ckan("datastore_search", resource_id=resource_id, limit=MAX_DATASTORE_ROWS)
    records = res.get("records") or []
    if not records:
        raise PullError("datastore_search returned 0 records for resource %s" % resource_id)
    headers = list(records[0].keys())
    rows = [headers]
    for rec in records:
        rows.append([("" if rec.get(h) is None else str(rec.get(h))).strip()
                     for h in headers])
    return rows


# ── generic table -> per-series monthly price map ────────────────────────────
def _norm(s):
    return re.sub(r"\s+", " ", (s or "").strip())


def _header_find(headers, keywords):
    for i, h in enumerate(headers):
        hl = _norm(h).lower()
        for kw in keywords:
            if kw in hl:
                return i
    return None


def _parse_year(v):
    """Year cell -> CE int; BE (>2200) converted; tolerates '2568/69' ranges."""
    m = re.search(r"(\d{4})", str(v))
    if not m:
        return None
    y = int(m.group(1))
    if y > 2200:  # Buddhist Era
        y -= 543
    if not (1990 <= y <= 2200):
        return None
    return y


def _parse_month(v):
    s = _norm(str(v))
    if not s:
        return None
    if s in THAI_MONTHS:
        return THAI_MONTHS[s]
    for name, num in THAI_MONTHS.items():
        if s.startswith(name):
            return num
    m = re.match(r"^(\d{1,2})$", s)
    if m and 1 <= int(m.group(1)) <= 12:
        return int(m.group(1))
    return None


def _parse_price(v):
    s = _norm(str(v)).replace(",", "")
    if not s or s in ("-", "–", "n/a", "N/A", "…", "."):
        return None
    try:
        x = float(s)
    except ValueError:
        return None
    return x if x > 0 else None


def _parse_date_cell(v):
    """'2569-05' / '05/2569' / '2026-05-01' / 'พ.ค. 2569' -> (ce_year, month)."""
    s = _norm(str(v))
    m = re.match(r"^(\d{4})[-/.](\d{1,2})(?:[-/.]\d{1,2})?$", s)
    if m:
        y, mo = _parse_year(m.group(1)), int(m.group(2))
        return (y, mo) if (y and 1 <= mo <= 12) else None
    m = re.match(r"^(\d{1,2})[-/.](\d{4})$", s)
    if m:
        y, mo = _parse_year(m.group(2)), int(m.group(1))
        return (y, mo) if (y and 1 <= mo <= 12) else None
    for name, num in THAI_MONTHS.items():
        if s.startswith(name):
            y = _parse_year(s[len(name):])
            if y:
                return (y, num)
    return None


def parse_table(rows):
    """rows (list of lists, first = header) -> (series, units)

    series: {label_th: {(ce_year, month): price}}   month=0 for annual-only data
    units:  {label_th: unit or None}
    Handles LONG layout (label/year/month|date/price columns) and WIDE layout
    (one row per label+year, one column per Thai month).
    """
    if len(rows) < 2:
        raise PullError("table has no data rows")
    headers = rows[0]
    li = _header_find(headers, H_LABEL)
    if li is None:
        raise PullError("could not find a commodity-label column; headers: %s"
                        % headers)
    yi = _header_find(headers, H_YEAR)
    mi = _header_find(headers, H_MONTH)
    di = _header_find(headers, H_DATE)
    ui = _header_find(headers, H_UNIT)

    # WIDE layout: month-named columns
    month_cols = [(i, _parse_month(h)) for i, h in enumerate(headers)
                  if _parse_month(h) and i != li]
    series, units = {}, {}

    def put(label, key, price, unit):
        if price is None or not label:
            return
        series.setdefault(label, {})[key] = price
        if unit and label not in units:
            units[label] = unit

    if month_cols and yi is not None:
        for r in rows[1:]:
            if len(r) <= li:
                continue
            label = _norm(r[li])
            year = _parse_year(r[yi]) if len(r) > yi else None
            unit = _norm(r[ui]) if (ui is not None and len(r) > ui) else None
            if not year:
                continue
            for ci, mo in month_cols:
                if len(r) > ci:
                    put(label, (year, mo), _parse_price(r[ci]), unit)
    else:
        pi = _header_find(headers, H_PRICE)
        if pi is None:
            raise PullError("could not find a price column; headers: %s" % headers)
        for r in rows[1:]:
            if len(r) <= max(li, pi):
                continue
            label = _norm(r[li])
            price = _parse_price(r[pi])
            unit = _norm(r[ui]) if (ui is not None and len(r) > ui) else None
            key = None
            if di is not None and len(r) > di:
                key = _parse_date_cell(r[di])
            if key is None and yi is not None and len(r) > yi:
                year = _parse_year(r[yi])
                if year:
                    mo = _parse_month(r[mi]) if (mi is not None and len(r) > mi) else None
                    key = (year, mo if mo else 0)   # 0 = annual granularity
            if key:
                put(label, key, price, unit)

    if not series:
        raise PullError("parsed 0 usable series rows — layout not recognized; "
                        "headers were: %s" % headers)
    return series, units


# ── crop matching + YoY ──────────────────────────────────────────────────────
def match_crops(series):
    """Match series labels to target crops by keyword priority; deterministic."""
    matched = {}
    labels = sorted(series.keys())
    for crop, patterns in CROP_PATTERNS.items():
        for pat in patterns:
            cands = [lb for lb in labels if pat in _norm(lb).replace(" ", "")
                     or pat in _norm(lb)]
            if cands:
                # deterministic pick: most data points, then shortest, then lexical
                cands.sort(key=lambda lb: (-len(series[lb]), len(lb), lb))
                matched[crop] = cands[0]
                break
    return matched


def fmt_period(key):
    y, m = key
    return "%04d-%02d" % (y, m) if m else "%04d" % y


def fmt_period_be(key):
    y, m = key
    return "%04d-%02d" % (y + 543, m) if m else "%04d" % (y + 543)


def yoy_for(points):
    """points: {(y,m): price} -> (latest_key, prior_key, yoy%) using the series
    itself: latest period vs the SAME period one year earlier."""
    keys = sorted(points.keys())
    latest = keys[-1]
    prior = (latest[0] - 1, latest[1])
    if prior in points and points[prior]:
        yoy = round((points[latest] / points[prior] - 1.0) * 100.0, 2)
        return latest, prior, yoy
    return latest, None, None


def build_landing(series, units, matched, provenance, stamp):
    commodities = {}
    vintage_key = None
    for crop in sorted(CROP_PATTERNS.keys()):
        label = matched.get(crop)
        if not label:
            continue
        pts = series[label]
        latest, prior, yoy = yoy_for(pts)
        if vintage_key is None or latest > vintage_key:
            vintage_key = latest
        entry = {
            "series_label_th": label,
            "unit": units.get(label),
            "latest": {"period": fmt_period(latest),
                       "period_be": fmt_period_be(latest),
                       "price": pts[latest]},
            "prior_year": ({"period": fmt_period(prior),
                            "period_be": fmt_period_be(prior),
                            "price": pts[prior]} if prior else None),
            "yoy": yoy,
            "n_points": len(pts),
        }
        commodities[crop] = entry

    usable = [c for c, e in commodities.items() if e["yoy"] is not None]
    if len(usable) < MIN_CROPS:
        raise PullError(
            "only %d/%d target crops got a usable YoY (%s) — refusing to land a "
            "thin file that would demote the honest World Bank proxy. Matched "
            "labels: %s" % (len(usable), len(CROP_PATTERNS), usable,
                            json.dumps(matched, ensure_ascii=False)))

    meta = {
        "label": "MEASURED — OAE farm-gate prices (ราคาที่เกษตรกรขายได้), prices "
                 "farmers actually received. Nothing modelled or estimated.",
        "generated_by": "pipeline/pull_oae_prices.py",
        "search_term": SEARCH_TERM,
        "vintage": "BE %s (CE %s)" % (fmt_period_be(vintage_key), fmt_period(vintage_key)),
        "vintage_ce": {"year": vintage_key[0], "month": vintage_key[1]},
        "pulled": stamp,
        "yoy_method": "computed from the series itself: latest period vs the same "
                      "period one year earlier (period granularity = monthly where "
                      "the source is monthly, annual otherwise).",
        "crops_covered": usable,
        "crops_missing": sorted(set(CROP_PATTERNS) - set(commodities)),
        "consumer": "pipeline/build_crop_stress.py prefers these measured YoY values "
                    "over the World Bank GLOBAL proxy for matching crops.",
    }
    meta.update(provenance)
    return {"meta": meta, "commodities": commodities}


# ── live pull ────────────────────────────────────────────────────────────────
def run_pull(out_path, stamp, dry_run):
    pkg = find_package()
    res = pick_resource(pkg)
    print("resource: %s  format=%s  url=%s"
          % (res.get("name") or res.get("id"), res.get("format"), res.get("url")))
    if dry_run:
        print("\n-- dry run: all resources of the chosen dataset --")
        for r in pkg.get("resources") or []:
            print("  [%s] %s  %s" % (r.get("format"), r.get("name") or r.get("id"),
                                     r.get("url")))
        print("dry run complete — nothing downloaded, nothing written.")
        return

    rank = rank_resource(res)
    if rank == 2:
        rows = rows_from_datastore(res["id"])
    else:
        raw = http_get(res["url"], timeout=180)
        rows = rows_from_csv(raw) if rank == 0 else rows_from_xlsx(raw)
    print("downloaded %d rows" % (len(rows) - 1))

    series, units = parse_table(rows)
    matched = match_crops(series)
    print("matched crops: %s" % json.dumps(
        {c: matched[c] for c in sorted(matched)}, ensure_ascii=False))
    provenance = {
        "source": BASE,
        "dataset_id": pkg.get("id"),
        "dataset_title": pkg.get("title"),
        "dataset_modified": pkg.get("metadata_modified"),
        "resource_id": res.get("id"),
        "resource_url": res.get("url"),
        "resource_format": res.get("format"),
    }
    doc = build_landing(series, units, matched, provenance, stamp)
    write_landing(doc, out_path)


def write_landing(doc, out_path):
    text = json.dumps(doc, ensure_ascii=False, indent=2) + "\n"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)
    print("wrote %s  (vintage %s, %d crops: %s)"
          % (out_path, doc["meta"]["vintage"], len(doc["commodities"]),
             ", ".join(sorted(doc["commodities"]))))
    for crop in sorted(doc["commodities"]):
        e = doc["commodities"][crop]
        print("  %-10s yoy=%-8s latest=%s %s  [%s]"
              % (crop, e["yoy"], e["latest"]["price"], e["unit"] or "",
                 e["series_label_th"]))


# ── offline selftest ─────────────────────────────────────────────────────────
# Fixture A: LONG layout — Thai headers, BE years, comma-grouped numbers, a
# decoy commodity, and a 13-month rice series (proves latest-vs-prior-year pick).
FIXTURE_LONG = """\
ชนิดสินค้า,ปี,เดือน,ราคา,หน่วย
ข้าวเปลือกเจ้าความชื้น 15%,2568,5,"8,100.50",บาท/ตัน
ข้าวเปลือกเจ้าความชื้น 15%,2568,6,"8,050.00",บาท/ตัน
ข้าวเปลือกเจ้าความชื้น 15%,2569,4,"7,600.00",บาท/ตัน
ข้าวเปลือกเจ้าความชื้น 15%,2569,5,"7,450.25",บาท/ตัน
ยางแผ่นดิบชั้น 3,2568,5,58.10,บาท/กก.
ยางแผ่นดิบชั้น 3,2569,5,49.75,บาท/กก.
อ้อยโรงงาน,2568,5,"1,180.00",บาท/ตัน
อ้อยโรงงาน,2569,5,"1,050.00",บาท/ตัน
ผลปาล์มน้ำมันทั้งทะลาย,2568,5,5.80,บาท/กก.
ผลปาล์มน้ำมันทั้งทะลาย,2569,5,4.95,บาท/กก.
หัวมันสำปะหลังสดคละ,2568,5,2.45,บาท/กก.
หัวมันสำปะหลังสดคละ,2569,5,1.90,บาท/กก.
ข้าวโพดเลี้ยงสัตว์ความชื้น 14.5%,2568,5,10.10,บาท/กก.
ข้าวโพดเลี้ยงสัตว์ความชื้น 14.5%,2569,5,10.35,บาท/กก.
กระเทียมแห้งใหญ่ คละ,2569,5,32.50,บาท/กก.
"""

# Fixture B: WIDE layout — one column per Thai month, BE year column.
FIXTURE_WIDE = """\
ชนิดสินค้า,ปี,ม.ค.,ก.พ.,มี.ค.,หน่วย
ข้าวเปลือกเจ้าความชื้น 15%,2568,"8,200.00","8,150.00","8,100.00",บาท/ตัน
ข้าวเปลือกเจ้าความชื้น 15%,2569,"7,700.00","7,650.00","7,600.00",บาท/ตัน
ยางแผ่นดิบชั้น 3,2568,60.00,59.00,58.00,บาท/กก.
ยางแผ่นดิบชั้น 3,2569,52.00,51.00,50.00,บาท/กก.
"""


def _expect(cond, what):
    if cond:
        print("  PASS  %s" % what)
        return 0
    print("  FAIL  %s" % what)
    return 1


def validate_landing(doc):
    """Schema validation shared by selftest (and usable ad-hoc). Returns #fails."""
    bad = 0
    meta = doc.get("meta") or {}
    for k in ("label", "generated_by", "vintage", "vintage_ce", "yoy_method",
              "crops_covered", "source", "dataset_id", "resource_url"):
        bad += _expect(k in meta, "meta.%s present" % k)
    bad += _expect(str(meta.get("label", "")).startswith("MEASURED"),
                   "meta.label starts with MEASURED")
    vc = meta.get("vintage_ce") or {}
    bad += _expect(isinstance(vc.get("year"), int) and vc["year"] >= 2020,
                   "meta.vintage_ce.year is a sane CE year (got %r)" % vc.get("year"))
    comms = doc.get("commodities") or {}
    bad += _expect(set(comms) <= set(CROP_PATTERNS),
                   "commodity keys are all known crops")
    for crop, e in sorted(comms.items()):
        bad += _expect(isinstance(e.get("latest", {}).get("price"), (int, float)),
                       "%s latest.price numeric" % crop)
        bad += _expect(e.get("yoy") is None or isinstance(e["yoy"], (int, float)),
                       "%s yoy numeric-or-null" % crop)
        bad += _expect(bool(e.get("series_label_th")), "%s has a Thai series label" % crop)
    return bad


def run_selftest():
    print("== selftest: LONG fixture through the live parse path ==")
    bad = 0
    series, units = parse_table(rows_from_csv(FIXTURE_LONG.encode("utf-8")))
    matched = match_crops(series)
    bad += _expect(set(matched) == set(CROP_PATTERNS),
                   "all 6 crops matched (got %s)" % sorted(matched))
    bad += _expect("กระเทียมแห้งใหญ่ คละ" not in matched.values(),
                   "decoy commodity (garlic) not matched to any crop")
    prov = {"source": BASE, "dataset_id": "SELFTEST", "dataset_title": "fixture",
            "dataset_modified": None, "resource_id": "SELFTEST",
            "resource_url": "embedded://fixture-long", "resource_format": "CSV"}
    doc = build_landing(series, units, matched, prov, stamp="1970-01-01")
    bad += validate_landing(doc)
    # exact YoY numbers, computed by hand from the fixture
    expect_yoy = {"rice": -8.03, "rubber": -14.37, "sugarcane": -11.02,
                  "oilpalm": -14.66, "cassava": -22.45, "maize": 2.48}
    for crop, want in sorted(expect_yoy.items()):
        got = doc["commodities"][crop]["yoy"]
        bad += _expect(got == want, "%s yoy == %s (got %s)" % (crop, want, got))
    bad += _expect(doc["meta"]["vintage_ce"] == {"year": 2026, "month": 5},
                   "vintage = latest period in the data (2026-05)")
    bad += _expect(doc["commodities"]["rice"]["latest"]["period_be"] == "2569-05",
                   "BE period preserved alongside CE")
    bad += _expect(doc["meta"]["pulled"] == "1970-01-01",
                   "pulled comes from --stamp, not wall clock")

    print("== selftest: WIDE (month-column) fixture ==")
    series2, units2 = parse_table(rows_from_csv(FIXTURE_WIDE.encode("utf-8")))
    matched2 = match_crops(series2)
    bad += _expect(set(matched2) == {"rice", "rubber"},
                   "wide fixture matches rice+rubber (got %s)" % sorted(matched2))
    doc2 = build_landing(series2, units2, matched2, prov, stamp="1970-01-01")
    got = doc2["commodities"]["rice"]["yoy"]
    bad += _expect(got == -6.17, "wide rice yoy == -6.17 (Mar/Mar) (got %s)" % got)
    bad += _expect(doc2["meta"]["vintage_ce"] == {"year": 2026, "month": 3},
                   "wide vintage = 2026-03")

    print("\nselftest: %s" % ("ALL PASS" if bad == 0 else "%d FAILURE(S)" % bad))
    return bad


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out", default=DEFAULT_OUT,
                    help="landing file path (default: %s)" % DEFAULT_OUT)
    ap.add_argument("--stamp", default=None,
                    help="pull date YYYY-MM-DD for meta.pulled — REQUIRED for a "
                         "real write; never wall-clock-embedded by the script")
    ap.add_argument("--dry-run", action="store_true",
                    help="search the catalog + list resources only; no write")
    ap.add_argument("--selftest", action="store_true",
                    help="OFFLINE: parse embedded fixtures through the live parse "
                         "path and validate the output schema + exact YoY numbers")
    args = ap.parse_args()

    if args.selftest:
        sys.exit(1 if run_selftest() else 0)

    if not args.dry_run:
        if not args.stamp:
            ap.error("--stamp YYYY-MM-DD is required for a real pull (provenance: "
                     "the pull date must be stated, and the script never embeds "
                     "the wall clock itself). Use --dry-run to only look.")
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", args.stamp):
            ap.error("--stamp must be YYYY-MM-DD, got %r" % args.stamp)

    try:
        run_pull(args.out, args.stamp, args.dry_run)
    except PullError as e:
        print("\nPULL FAILED: %s" % e, file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
